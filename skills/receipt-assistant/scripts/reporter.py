"""报表生成器 - 生成汇总 Excel"""
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

try:
    from .models import TicketInfo, Summary
except ImportError:
    from models import TicketInfo, Summary


class ReportGenerator:
    """报表生成器"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate_excel(self, tickets: List[TicketInfo], filename: str = "报销汇总.xlsx") -> str:
        """生成 Excel 汇总表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "报销汇总"

        # 设置表头
        headers = [
            "序号", "票据类型", "原始文件名", "新文件名",
            "开票日期", "发票号码", "金额", "销售方", "购买方", "乘客/旅客", "状态"
        ]
        self._write_header(ws, headers)

        # 写入数据
        for idx, ticket in enumerate(tickets, start=1):
            row = [
                idx,
                ticket.file_type.value,
                ticket.original_name,
                ticket.new_name,
                ticket.date,
                ticket.invoice_no,
                ticket.amount,
                ticket.seller,
                ticket.buyer,
                ticket.person,
                ticket.status.value,
            ]
            ws.append(row)

        # 设置列宽
        self._set_column_width(ws)

        # 保存文件
        output_path = self.output_dir / filename
        wb.save(output_path)

        return str(output_path)

    def _write_header(self, ws, headers: List[str]):
        """写入表头"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _set_column_width(self, ws):
        """设置列宽"""
        widths = [8, 10, 40, 40, 12, 20, 12, 25, 25, 12, 10]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    def generate_summary(self, tickets: List[TicketInfo]) -> Summary:
        """生成汇总统计"""
        summary = Summary()

        for ticket in tickets:
            summary.add_ticket(ticket)

        return summary

    def format_summary(self, summary: Summary) -> str:
        """格式化汇总信息为字符串"""
        lines = [
            "=" * 40,
            "报销汇总",
            "=" * 40,
            f"总文件数: {summary.total_count}",
            f"成功: {summary.success_count}",
            f"待确认: {summary.pending_count}",
            f"失败: {summary.error_count}",
            f"总金额: {summary.total_amount:.2f} 元",
            "-" * 40,
            "分类统计:",
        ]

        for type_name, count in summary.type_counts.items():
            lines.append(f"  {type_name}: {count} 个")

        lines.append("=" * 40)

        return "\n".join(lines)
