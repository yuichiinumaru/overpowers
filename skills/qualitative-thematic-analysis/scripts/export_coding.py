#!/usr/bin/env python3
"""
export_coding.py — 将编码结果导出为 Excel 文件
用法：python3 export_coding.py <编码JSON路径> <输出Excel路径>

编码JSON格式：
[
  {
    "code": "编码名称",
    "sub_code": "子编码",
    "quote": "原文片段",
    "source": "来源文件名",
    "line": "行号/段落号",
    "theme": "所属主题",
    "note": "备注"
  }
]
"""
import sys
import json
import os

def export_to_excel(codes, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("需要安装 openpyxl：pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "编码表"

    # 表头
    headers = ["主题", "编码", "子编码", "原文片段", "来源文件", "位置", "备注"]
    header_fill = PatternFill("solid", fgColor="2D3561")
    header_font = Font(bold=True, color="FFFFFF", name="微软雅黑", size=11)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # 列宽
    col_widths = [20, 20, 20, 50, 20, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 30

    # 数据行
    alt_fill = PatternFill("solid", fgColor="F5F7FF")
    for row_idx, code in enumerate(codes, 2):
        row_fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            code.get("theme", ""),
            code.get("code", ""),
            code.get("sub_code", ""),
            code.get("quote", ""),
            code.get("source", ""),
            code.get("line", ""),
            code.get("note", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"✅ 编码表已保存：{output_path}（共 {len(codes)} 条编码）")

def main():
    if len(sys.argv) < 3:
        print("用法：python3 export_coding.py <编码JSON路径> <输出Excel路径>")
        sys.exit(1)

    json_path = os.path.expanduser(sys.argv[1])
    output_path = os.path.expanduser(sys.argv[2])

    with open(json_path, 'r', encoding='utf-8') as f:
        codes = json.load(f)

    export_to_excel(codes, output_path)

if __name__ == "__main__":
    main()
