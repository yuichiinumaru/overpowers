#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Markdown 到 Excel 转换器
支持将 Markdown 表格、列表转换为 Excel 工作表
"""

import re
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class MarkdownToExcelConverter:
    """Markdown 到 Excel 转换器"""
    
    def __init__(self):
        self.wb = None
        self.default_font = '微软雅黑'
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_font = Font(name=self.default_font, size=11, bold=True, color='FFFFFF')
        self.cell_font = Font(name=self.default_font, size=10)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def convert(self, md_content: str, output_path: str, title: str = None):
        """
        将 Markdown 内容转换为 Excel 文件
        
        Args:
            md_content: Markdown 文本内容
            output_path: 输出 Excel 文件路径
            title: 文档标题（可选，作为第一个工作表名称）
        """
        self.wb = Workbook()
        
        # 删除默认工作表
        if 'Sheet' in self.wb.sheetnames:
            self.wb.remove(self.wb['Sheet'])
        
        # 解析 Markdown 并创建表格
        self._parse_markdown(md_content, title)
        
        # 如果没有创建任何工作表，创建一个空的
        if len(self.wb.sheetnames) == 0:
            self.wb.create_sheet(title='Sheet1')
        
        # 保存文件
        self.wb.save(output_path)
        return output_path
    
    def convert_file(self, md_file: str, output_path: str = None, title: str = None):
        """
        将 Markdown 文件转换为 Excel 文件
        
        Args:
            md_file: Markdown 文件路径
            output_path: 输出 Excel 文件路径（可选）
            title: 文档标题（可选）
        """
        # 读取 Markdown 文件
        with open(md_file, 'r', encoding='utf-8') as f:
            md_content = f.read()
        
        # 确定输出路径
        if output_path is None:
            base_name = os.path.splitext(md_file)[0]
            output_path = base_name + '.xlsx'
        
        return self.convert(md_content, output_path, title)
    
    def _parse_markdown(self, md_content: str, default_title: str = None):
        """解析 Markdown 内容"""
        lines = md_content.split('\n')
        i = 0
        sheet_index = 0
        current_title = default_title or 'Sheet1'
        
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            
            # 空行
            if not stripped:
                i += 1
                continue
            
            # 标题 - 可能作为工作表名称
            if stripped.startswith('#'):
                title_match = re.match(r'^#+\s+(.+)$', stripped)
                if title_match:
                    current_title = title_match.group(1)[:31]  # Excel 工作表名称最多31个字符
                i += 1
                continue
            
            # 表格
            if '|' in stripped and i + 1 < len(lines) and '|' in lines[i + 1] and '---' in lines[i + 1]:
                i = self._handle_table(lines, i, current_title, sheet_index)
                sheet_index += 1
                continue
            
            # 列表 - 转换为单列数据
            if stripped.startswith(('- ', '* ', '+ ')) or re.match(r'^\d+\.', stripped):
                i = self._handle_list(lines, i, current_title, sheet_index)
                sheet_index += 1
                continue
            
            i += 1
    
    def _handle_table(self, lines: list, start_idx: int, sheet_name: str, sheet_index: int) -> int:
        """处理表格"""
        # 解析表头
        header_line = lines[start_idx].strip()
        headers = [cell.strip() for cell in header_line.split('|') if cell.strip()]
        
        # 跳过分隔行
        i = start_idx + 2
        
        # 收集数据行
        rows = []
        while i < len(lines) and '|' in lines[i]:
            row_line = lines[i].strip()
            cells = [cell.strip() for cell in row_line.split('|') if cell.strip()]
            if cells:
                rows.append(cells)
            i += 1
        
        # 创建工作表
        ws_name = f"{sheet_name[:20]}_{sheet_index}" if sheet_index > 0 else sheet_name[:31]
        ws = self.wb.create_sheet(title=ws_name)
        
        # 写入表头
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.border
            
            # 写入数据
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    if col_idx <= len(headers):
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        cell.font = self.cell_font
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = self.border
            
            # 调整列宽
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 20
        
        return i
    
    def _handle_list(self, lines: list, start_idx: int, sheet_name: str, sheet_index: int) -> int:
        """处理列表"""
        i = start_idx
        items = []
        is_ordered = False
        
        while i < len(lines):
            line = lines[i]
            stripped = line.strip()
            
            # 无序列表
            if stripped.startswith(('- ', '* ', '+ ')):
                content = stripped[2:]
                items.append(content)
                i += 1
            # 有序列表
            elif re.match(r'^\d+\.', stripped):
                content = re.sub(r'^\d+\.\s*', '', stripped)
                items.append(content)
                is_ordered = True
                i += 1
            else:
                break
        
        # 创建工作表
        list_type = "有序列表" if is_ordered else "无序列表"
        ws_name = f"{sheet_name[:15]}_{list_type}_{sheet_index}"[:31]
        ws = self.wb.create_sheet(title=ws_name)
        
        # 写入表头
        header_cell = ws.cell(row=1, column=1, value="序号" if is_ordered else "项目")
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = self.border
        
        # 写入数据
        for row_idx, item in enumerate(items, 2):
            cell = ws.cell(row=row_idx, column=1, value=item)
            cell.font = self.cell_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.border
        
        # 调整列宽
        ws.column_dimensions['A'].width = 50
        
        return i


class ExcelToMarkdownConverter:
    """Excel 到 Markdown 转换器"""
    
    def __init__(self):
        pass
    
    def convert(self, xlsx_path: str, output_path: str = None) -> str:
        """
        将 Excel 文件转换为 Markdown
        
        Args:
            xlsx_path: Excel 文件路径
            output_path: 输出 Markdown 文件路径（可选）
        
        Returns:
            Markdown 文本内容
        """
        from openpyxl import load_workbook
        
        wb = load_workbook(xlsx_path)
        md_content = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 添加工作表标题
            md_content.append(f"## {sheet_name}\n")
            
            # 转换为表格
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                # 表头
                header = rows[0]
                md_content.append('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in header) + ' |')
                
                # 分隔符
                md_content.append('|' + '|'.join([' --- ' for _ in header]) + '|')
                
                # 数据行
                for row in rows[1:]:
                    md_content.append('| ' + ' | '.join(str(cell) if cell is not None else '' for cell in row) + ' |')
                
                md_content.append('')
        
        md_text = '\n'.join(md_content)
        
        # 保存到文件
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(md_text)
        
        return md_text


class ExcelImageHandler:
    """Excel 图片处理器 - 支持图文混排"""
    
    def __init__(self, wb=None):
        self.wb = wb
        
    def set_workbook(self, wb):
        """设置工作簿对象"""
        self.wb = wb
    
    def insert_image_to_cell(self, sheet_name: str, cell: str, image_path: str,
                             width: float = None, height: float = None) -> bool:
        """
        在指定单元格插入图片
        
        Args:
            sheet_name: 工作表名称
            cell: 单元格位置（如 'A1'）
            image_path: 图片路径
            width: 图片宽度（像素，可选）
            height: 图片高度（像素，可选）
        
        Returns:
            是否成功
        """
        try:
            if not os.path.exists(image_path):
                print(f"图片不存在: {image_path}")
                return False
            
            if sheet_name not in self.wb.sheetnames:
                print(f"工作表不存在: {sheet_name}")
                return False
            
            ws = self.wb[sheet_name]
            
            # 插入图片
            img = openpyxl.drawing.image.Image(image_path)
            
            # 设置尺寸
            if width:
                img.width = width
            if height:
                img.height = height
            
            # 将图片添加到单元格
            ws.add_image(img, cell)
            
            return True
            
        except Exception as e:
            print(f"插入图片失败: {e}")
            return False
    
    def insert_floating_image(self, sheet_name: str, image_path: str,
                              left: float = 0, top: float = 0,
                              width: float = None, height: float = None) -> bool:
        """
        在工作表插入浮动图片
        
        Args:
            sheet_name: 工作表名称
            image_path: 图片路径
            left: 左边距（像素）
            top: 上边距（像素）
            width: 图片宽度（像素，可选）
            height: 图片高度（像素，可选）
        
        Returns:
            是否成功
        """
        try:
            if not os.path.exists(image_path):
                print(f"图片不存在: {image_path}")
                return False
            
            if sheet_name not in self.wb.sheetnames:
                print(f"工作表不存在: {sheet_name}")
                return False
            
            ws = self.wb[sheet_name]
            
            # 插入图片
            img = openpyxl.drawing.image.Image(image_path)
            
            # 设置尺寸
            if width:
                img.width = width
            if height:
                img.height = height
            
            # 设置位置
            from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            img.anchor = openpyxl.drawing.spreadsheet_drawing.OneCellAnchor()
            img.anchor._from = XDRPoint2D(pixels_to_EMU(left), pixels_to_EMU(top))
            
            ws.add_image(img)
            
            return True
            
        except Exception as e:
            print(f"插入浮动图片失败: {e}")
            return False
    
    def create_text_image_layout(self, sheet_name: str, cell: str, text: str,
                                  image_path: str, layout: str = 'right',
                                  image_width: float = 100) -> bool:
        """
        在单元格区域创建图文混排布局
        
        Args:
            sheet_name: 工作表名称
            cell: 起始单元格位置（如 'A1'）
            text: 文本内容
            image_path: 图片路径
            layout: 布局方式 (left/right/top/bottom)
            image_width: 图片宽度（像素）
        
        Returns:
            是否成功
        """
        try:
            if not os.path.exists(image_path):
                print(f"图片不存在: {image_path}")
                return False
            
            if sheet_name not in self.wb.sheetnames:
                print(f"工作表不存在: {sheet_name}")
                return False
            
            ws = self.wb[sheet_name]
            
            # 解析起始单元格
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            col_letter, row = coordinate_from_string(cell)
            col = column_index_from_string(col_letter)
            
            if layout == 'left':
                # 图片在左，文字在右
                img_cell = cell
                text_cell = f"{get_column_letter(col + 3)}{row}"  # 图片占2列，间隔1列
            elif layout == 'right':
                # 图片在右，文字在左
                text_cell = cell
                img_cell = f"{get_column_letter(col + 3)}{row}"
            elif layout == 'top':
                # 图片在上，文字在下
                img_cell = cell
                text_cell = f"{col_letter}{row + 5}"  # 图片占4行
            elif layout == 'bottom':
                # 文字在上，图片在下
                text_cell = cell
                img_cell = f"{col_letter}{row + 2}"
            else:
                print(f"不支持的布局方式: {layout}")
                return False
            
            # 添加文字
            ws[text_cell] = text
            ws[text_cell].alignment = Alignment(wrap_text=True, vertical='top')
            
            # 插入图片
            img = openpyxl.drawing.image.Image(image_path)
            img.width = image_width
            img.height = image_width  # 保持正方形
            
            ws.add_image(img, img_cell)
            
            return True
            
        except Exception as e:
            print(f"创建图文混排失败: {e}")
            return False


# 便捷函数
def md_to_excel(md_file: str, output_path: str = None, title: str = None) -> str:
    """Markdown 文件转 Excel"""
    converter = MarkdownToExcelConverter()
    return converter.convert_file(md_file, output_path, title)


def excel_to_md(xlsx_file: str, output_path: str = None) -> str:
    """Excel 文件转 Markdown"""
    converter = ExcelToMarkdownConverter()
    return converter.convert(xlsx_file, output_path)


def insert_image_to_excel(xlsx_file: str, sheet_name: str, cell: str,
                          image_path: str, output_path: str = None,
                          width: float = None, height: float = None) -> str:
    """
    向 Excel 指定单元格插入图片
    
    Args:
        xlsx_file: Excel 文件路径
        sheet_name: 工作表名称
        cell: 单元格位置（如 'A1'）
        image_path: 图片路径
        output_path: 输出路径（可选）
        width: 图片宽度（像素，可选）
        height: 图片高度（像素，可选）
    
    Returns:
        输出文件路径
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(xlsx_file)
    handler = ExcelImageHandler(wb)
    
    handler.insert_image_to_cell(sheet_name, cell, image_path, width, height)
    
    if output_path is None:
        output_path = xlsx_file
    
    wb.save(output_path)
    return output_path


def create_excel_text_image_layout(xlsx_file: str, sheet_name: str, cell: str,
                                    text: str, image_path: str, layout: str = 'right',
                                    image_width: float = 100, output_path: str = None) -> str:
    """
    在 Excel 创建图文混排布局
    
    Args:
        xlsx_file: Excel 文件路径
        sheet_name: 工作表名称
        cell: 起始单元格位置（如 'A1'）
        text: 文本内容
        image_path: 图片路径
        layout: 布局方式 (left/right/top/bottom)
        image_width: 图片宽度（像素）
        output_path: 输出路径（可选）
    
    Returns:
        输出文件路径
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(xlsx_file)
    handler = ExcelImageHandler(wb)
    
    handler.create_text_image_layout(sheet_name, cell, text, image_path, layout, image_width)
    
    if output_path is None:
        output_path = xlsx_file
    
    wb.save(output_path)
    return output_path


if __name__ == '__main__':
    # 测试代码
    test_md = """# 销售数据

## 产品销量

| 产品 | 一月 | 二月 | 三月 |
|------|------|------|------|
| 产品A | 100 | 150 | 200 |
| 产品B | 80 | 120 | 160 |

## 地区分布

| 地区 | 销售额 |
|------|--------|
| 北京 | 50000 |
| 上海 | 45000 |

## 待办事项

- 完成季度报告
- 更新产品文档
- 安排团队会议

## 优先级

1. 紧急任务
2. 重要任务
3. 一般任务
"""
    
    converter = MarkdownToExcelConverter()
    output = '/tmp/test_excel.xlsx'
    converter.convert(test_md, output, '销售数据')
    print(f'转换完成: {output}')
