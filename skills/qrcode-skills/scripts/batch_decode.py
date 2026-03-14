"""
批量解码二维码 - 从 Excel/CSV/TXT 读取图片 URL，批量解码并写回结果。

用法:
  python scripts/batch_decode.py --input <文件> [选项]

选项:
  --column <列名或索引>   指定 URL 列（Excel/CSV 有效），不指定时自动检测
  --output-txt <文件>     单独输出结果到 TXT（按行分隔）

默认行为: 在原文件中新增一列写入解码结果。
TXT 输入时结果默认写到 <原文件名>_decoded.txt。

输出 JSON:
  成功: {"total": N, "success": N, "failed": N, "output_file": "...", "output_txt": "..."|null}
  需要用户指定列: {"need_column": true, "columns": [...], "preview": [...]}
  错误: {"error": "..."}
"""

import sys
import json
import os
import csv
import argparse
import tempfile
from pathlib import Path
from urllib.parse import quote

API_ENDPOINT = "https://api.2dcode.biz/v1/read-qr-code"
FAIL_PLACEHOLDER = "未解析到二维码"


def decode_single(image_source: str) -> str:
    """解码单个图片，返回内容文本或失败占位符。"""
    result = _try_zxing(image_source)
    if result:
        return result

    if _is_url(image_source):
        result = _decode_api_url(image_source)
    elif os.path.isfile(image_source):
        result = _decode_api_file(image_source)
    else:
        result = _decode_api_url(image_source)

    return result if result else FAIL_PLACEHOLDER


def _is_url(s: str) -> bool:
    return s.startswith("http://") or s.startswith("https://")


def _try_zxing(source: str) -> str | None:
    try:
        import zxingcpp
        from PIL import Image
    except ImportError:
        return None

    tmp_path = None
    try:
        if _is_url(source):
            import urllib.request
            suffix = Path(source.split("?")[0]).suffix or ".png"
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            urllib.request.urlretrieve(source, tmp.name)
            tmp.close()
            tmp_path = tmp.name
            img_path = tmp_path
        else:
            img_path = source

        img = Image.open(img_path)
        results = zxingcpp.read_barcodes(img)
        if results:
            return "; ".join(r.text for r in results)
        return None
    except Exception:
        return None
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _decode_api_url(image_url: str) -> str | None:
    import urllib.request

    api_url = f"{API_ENDPOINT}?file_url={quote(image_url, safe='')}"
    try:
        with urllib.request.urlopen(api_url) as resp:
            data = json.loads(resp.read().decode())
        contents = data.get("data", {}).get("contents", [])
        if data.get("code") == 0 and contents:
            return "; ".join(contents)
        return None
    except Exception:
        return None


def _decode_api_file(file_path: str) -> str | None:
    import urllib.request
    import mimetypes
    import uuid

    boundary = uuid.uuid4().hex
    filename = os.path.basename(file_path)
    mime_type = mimetypes.guess_type(file_path)[0] or "application/octet-stream"

    with open(file_path, "rb") as f:
        file_data = f.read()

    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        f"Content-Type: {mime_type}\r\n\r\n"
    ).encode() + file_data + f"\r\n--{boundary}--\r\n".encode()

    req = urllib.request.Request(
        API_ENDPOINT,
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode())
        contents = data.get("data", {}).get("contents", [])
        if data.get("code") == 0 and contents:
            return "; ".join(contents)
        return None
    except Exception:
        return None


# ── 文件读写 ──────────────────────────────────────────────

def process_txt(input_path: str, output_txt: str | None):
    with open(input_path, "r", encoding="utf-8") as f:
        lines = [line.strip() for line in f if line.strip()]

    results = []
    success = 0
    for line in lines:
        decoded = decode_single(line)
        results.append(decoded)
        if decoded != FAIL_PLACEHOLDER:
            success += 1

    out_path = output_txt or str(Path(input_path).with_stem(Path(input_path).stem + "_decoded").with_suffix(".txt"))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(results))

    return {
        "total": len(lines),
        "success": success,
        "failed": len(lines) - success,
        "output_file": os.path.abspath(out_path),
        "output_txt": os.path.abspath(out_path),
    }


def process_csv(input_path: str, column, output_txt: str | None):
    with open(input_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        return {"error": "CSV 文件为空"}

    headers = rows[0]
    data_rows = rows[1:]
    col_idx = _resolve_col(headers, column)

    if col_idx is None:
        return {
            "need_column": True,
            "columns": headers,
            "preview": rows[:6],
            "message": "无法自动判断 URL 列，请指定 --column 参数" if column is None else f"找不到列 '{column}'",
        }

    decoded_results = []
    success = 0
    for row in data_rows:
        url = row[col_idx].strip() if col_idx < len(row) else ""
        if url:
            decoded = decode_single(url)
        else:
            decoded = FAIL_PLACEHOLDER
        decoded_results.append(decoded)
        if decoded != FAIL_PLACEHOLDER:
            success += 1

    if output_txt:
        with open(output_txt, "w", encoding="utf-8") as f:
            f.write("\n".join(decoded_results))
        return {
            "total": len(data_rows),
            "success": success,
            "failed": len(data_rows) - success,
            "output_file": os.path.abspath(input_path),
            "output_txt": os.path.abspath(output_txt),
        }

    result_col = "解码结果"
    headers.append(result_col)
    for i, row in enumerate(data_rows):
        row.append(decoded_results[i])

    with open(input_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data_rows)

    return {
        "total": len(data_rows),
        "success": success,
        "failed": len(data_rows) - success,
        "output_file": os.path.abspath(input_path),
        "output_txt": None,
    }


def process_excel(input_path: str, column, output_txt: str | None):
    import openpyxl

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    col_idx = _resolve_col(headers, column)

    if col_idx is None:
        preview = []
        for row in ws.iter_rows(min_row=1, max_row=min(6, ws.max_row), values_only=True):
            preview.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return {
            "need_column": True,
            "columns": headers,
            "preview": preview,
            "message": "无法自动判断 URL 列，请指定 --column 参数" if column is None else f"找不到列 '{column}'",
        }

    decoded_results = []
    success = 0
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=col_idx + 1).value
        url = str(cell_val).strip() if cell_val else ""
        if url:
            decoded = decode_single(url)
        else:
            decoded = FAIL_PLACEHOLDER
        decoded_results.append(decoded)
        if decoded != FAIL_PLACEHOLDER:
            success += 1

    if output_txt:
        with open(output_txt, "w", encoding="utf-8") as f:
            f.write("\n".join(decoded_results))
        wb.close()
        return {
            "total": len(decoded_results),
            "success": success,
            "failed": len(decoded_results) - success,
            "output_file": os.path.abspath(input_path),
            "output_txt": os.path.abspath(output_txt),
        }

    result_col_idx = ws.max_column + 1
    ws.cell(row=1, column=result_col_idx, value="解码结果")
    for i, decoded in enumerate(decoded_results):
        ws.cell(row=i + 2, column=result_col_idx, value=decoded)

    wb.save(input_path)
    wb.close()

    return {
        "total": len(decoded_results),
        "success": success,
        "failed": len(decoded_results) - success,
        "output_file": os.path.abspath(input_path),
        "output_txt": None,
    }


# ── 列检测 ──────────────────────────────────────────────

def _resolve_col(headers: list[str], column) -> int | None:
    if column is not None:
        try:
            idx = int(column)
            return idx if 0 <= idx < len(headers) else None
        except ValueError:
            pass
        col_lower = str(column).lower().strip()
        for i, h in enumerate(headers):
            if h.strip().lower() == col_lower:
                return i
        return None

    keywords = ["url", "link", "image", "img", "图片", "链接", "网址", "二维码"]
    for i, h in enumerate(headers):
        h_lower = h.strip().lower()
        for kw in keywords:
            if kw in h_lower:
                return i
    if len(headers) == 1:
        return 0
    return None


# ── 入口 ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="输入文件路径")
    parser.add_argument("--column", default=None, help="指定 URL 列（列名或索引）")
    parser.add_argument("--output-txt", default=None, help="单独输出结果到 TXT")
    args = parser.parse_args()

    input_path = args.input
    ext = Path(input_path).suffix.lower()

    if ext == ".txt":
        result = process_txt(input_path, args.output_txt)
    elif ext == ".csv":
        result = process_csv(input_path, args.column, args.output_txt)
    elif ext in (".xlsx", ".xls"):
        result = process_excel(input_path, args.column, args.output_txt)
    else:
        result = {"error": f"不支持的文件格式: {ext}，支持 txt/csv/xlsx"}

    print(json.dumps(result, ensure_ascii=False))
    if "error" in result:
        sys.exit(1)


if __name__ == "__main__":
    main()
