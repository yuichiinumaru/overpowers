#!/usr/bin/env python3
"""財務データを Excel に出力する.

Usage:
    python3 export_excel.py \
      --company "株式会社サンプル" \
      --bs '{"total_assets": 300000000, "cash_and_deposits": 150000000, ...}' \
      --pl '{"revenue": 500000000, "operating_income": 50000000, ...}' \
      --valuation '{"suggested_price": 500000000, "median_per": 15.0, ...}' \
      --output ./output/report.xlsx

Output:
    Excel ファイルのパスを stdout に出力。
    3シート構成: B/S, P/L, Valuation
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, numbers
except ImportError:
    print("ERROR: openpyxl が必要です。`uv pip install openpyxl` を実行してください。", file=sys.stderr)
    sys.exit(1)

# 日本語ラベルマッピング
BS_LABELS = {
    "fiscal_date": "決算日",
    "cash_and_deposits": "現預金",
    "accounts_receivable": "売掛金",
    "inventory": "棚卸資産",
    "total_current_assets": "流動資産合計",
    "fixed_assets": "固定資産合計",
    "total_assets": "資産合計",
    "accounts_payable": "買掛金",
    "short_term_debt": "短期借入金",
    "officer_loans": "役員借入金",
    "total_current_liabilities": "流動負債合計",
    "long_term_debt": "長期借入金",
    "total_liabilities": "負債合計",
    "capital_stock": "資本金",
    "retained_earnings": "利益剰余金",
    "total_net_assets": "純資産合計",
}

PL_LABELS = {
    "fiscal_date": "決算日",
    "revenue": "売上高",
    "cost_of_sales": "売上原価",
    "gross_profit": "売上総利益",
    "sga_expenses": "販管費",
    "officer_compensation": "役員報酬",
    "operating_income": "営業利益",
    "non_operating_income": "営業外収益",
    "non_operating_expenses": "営業外費用",
    "ordinary_income": "経常利益",
    "net_income": "当期純利益",
    "depreciation": "減価償却費",
}

VALUATION_LABELS = {
    "method": "評価手法",
    "median_per": "中央値 PER",
    "median_ev_ebitda": "中央値 EV/EBITDA",
    "estimated_per_based": "PER ベース推定額",
    "estimated_ebitda_based": "EV/EBITDA ベース推定額",
    "discount_rate": "ディスカウント率",
    "suggested_price": "想定買収価格",
    "buy_signal": "買いシグナル",
    "ebitda": "EBITDA",
    "net_debt": "純有利子負債",
    "equity_ratio": "自己資本比率",
    "operating_margin": "売上高営業利益率",
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
MONEY_FORMAT = '#,##0'


def write_sheet(wb: Workbook, title: str, data: dict, labels: dict) -> None:
    """データを1シートに書き込む."""
    ws = wb.create_sheet(title=title)

    # ヘッダー
    for col, header in enumerate(["項目", "値"], start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # データ行
    for row_idx, (key, value) in enumerate(data.items(), start=2):
        label = labels.get(key, key)
        ws.cell(row=row_idx, column=1, value=label)

        cell = ws.cell(row=row_idx, column=2)
        if isinstance(value, (int, float)) and key != "fiscal_date":
            cell.value = value
            if isinstance(value, int) and value > 1000:
                cell.number_format = MONEY_FORMAT
        else:
            cell.value = str(value)

    # 列幅
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 25


def main() -> None:
    parser = argparse.ArgumentParser(description="財務データ Excel 出力")
    parser.add_argument("--company", required=True, help="会社名")
    parser.add_argument("--bs", help="B/S JSON 文字列")
    parser.add_argument("--pl", help="P/L JSON 文字列")
    parser.add_argument("--valuation", help="バリュエーション JSON 文字列")
    parser.add_argument("--output", required=True, help="出力 Excel ファイルパス")
    args = parser.parse_args()

    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    if args.bs:
        bs_data = json.loads(args.bs)
        write_sheet(wb, "BS 貸借対照表", bs_data, BS_LABELS)

    if args.pl:
        pl_data = json.loads(args.pl)
        write_sheet(wb, "PL 損益計算書", pl_data, PL_LABELS)

    if args.valuation:
        val_data = json.loads(args.valuation)
        write_sheet(wb, "Valuation", val_data, VALUATION_LABELS)

    if not wb.sheetnames:
        print("ERROR: --bs, --pl, --valuation のいずれかを指定してください。", file=sys.stderr)
        sys.exit(1)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"✓ Excel 出力: {output_path}", file=sys.stderr)
    print(str(output_path))


if __name__ == "__main__":
    main()
