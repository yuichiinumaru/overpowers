#!/usr/bin/env python3
"""
CRS 金融账户信息申报表生成器 v2.0
自动从PDF提取信息，生成美化版Excel表格
"""

import os
import sys
import re
from datetime import datetime

# ============== 配置 ==============
# 默认模板信息（脱敏）
DEFAULT_INFO = {
    "client_name": "[客户姓名]",
    "account_number": "[账户号码]",
    "account_type": "[账户类型]",
    "address": "[客户地址]",
    "tin": "[纳税人识别号]",
    "financial_institution": "[金融机构名称]",
    "financial_institution_id": "[金融机构ID]",
    
    # 资产数据（示例格式）
    "stock_value": "[股票市值]",
    "fund_value": "[基金市值]",
    "cash_value": "[现金余额]",
    "total_value": "[资产净值]",
    
    # 交易记录
    "transactions": [
        {
            "date": "[日期]",
            "type": "[买/卖/申购/赎回]",
            "name": "[产品名称]",
            "quantity": "[数量]",
            "amount": "[金额]"
        },
    ],
    
    # 费用
    "total_fees": "[费用合计]"
}

# ============== PDF解析 ==============
def parse_pdf(pdf_path):
    """从PDF提取信息"""
    try:
        import pdfplumber
        text = ""
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
        return extract_info_from_text(text)
    except Exception as e:
        print(f"PDF解析失败: {e}")
        return None

def extract_info_from_text(text):
    """从文本中提取CRS相关信息"""
    info = DEFAULT_INFO.copy()
    info["transactions"] = []
    
    # 提取客户姓名
    patterns = {
        "client_name": r"客户姓名[：:]\s*([^\n\s]+)|名称[：:]\s*([^\n\s]+)|Name[：:]\s*([^\n\s]+)",
        "account_number": r"账户号码[：:]\s*([0-9A-Z]{8,20})|帳戶號碼[：:]\s*([0-9A-Z]{8,20})",
        "account_type": r"账户类型[：:]\s*([^\n]+)|帳戶類型[：:]\s*([^\n]+)",
    }
    
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if match:
            for g in match.groups():
                if g:
                    info[key] = g.strip()
                    break
    
    # 提取金额
    amount_pattern = r"([0-9,]+(?:\.[0-9]+)?)\s*(HKD|USD|CNH|JPY|SGD)"
    amounts = re.findall(amount_pattern, text)
    if amounts:
        info["amounts_detected"] = [f"{a[0]} {a[1]}" for a in amounts[:10]]
    
    # 提取交易记录
    trade_patterns = [
        r"(买|卖|申购|赎回|申购|卖出)[^\n]{5,50}",
    ]
    for pattern in trade_patterns:
        matches = re.findall(pattern, text)
        for m in matches[:5]:
            if len(m) > 5:
                info["transactions"].append({
                    "date": "[日期]",
                    "type": "[交易类型]",
                    "name": m[:30],
                    "quantity": "[数量]",
                    "amount": "[金额]"
                })
    
    return info

# ============== Excel生成 ==============
def generate_excel(info, output_path="CRS_申报表.xlsx"):
    """生成美化版Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("请安装: pip install openpyxl")
        return None
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CRS申报表"
    
    # 样式定义
    colors = {
        "header": "1F4E79",
        "subheader": "2E75B6",
        "highlight": "FFE699",
        "light_gray": "D6DCE4"
    }
    
    def style_header(cell, text):
        cell.value = text
        cell.font = Font(name='微软雅黑', size=12, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors["header"], end_color=colors["header"], fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def style_subheader(cell, text):
        cell.value = text
        cell.font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors["subheader"], end_color=colors["subheader"], fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def style_cell(cell, text, bold=False, fill_color=None):
        cell.value = text
        cell.font = Font(name='微软雅黑', size=10, bold=bold)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def add_border(cell):
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 标题
    ws.merge_cells('A1:E1')
    ws['A1'] = "CRS 金融账户信息申报表"
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True, color=colors["header"])
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws['A2'] = f"申报期间: {datetime.now().strftime('%Y年%m月')}"
    ws['A2'].font = Font(name='微软雅黑', size=10, italic=True)
    ws.merge_cells('A2:E2')
    
    row = 4
    
    # 一、客户信息
    style_header(ws[f'A{row}'], "一、账户持有人信息")
    ws.merge_cells(f'A{row}:E{row}')
    ws.row_dimensions[row].height = 25
    row += 1
    
    for field, key in [("客户姓名", "client_name"), ("账户号码", "account_number"), 
                       ("账户类型", "account_type"), ("地址", "address"),
                       ("纳税人识别号", "tin")]:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = info.get(key, "[待补充]")
        style_cell(ws[f'A{row}'], field, bold=True, fill_color=colors["light_gray"])
        style_cell(ws[f'B{row}'], info.get(key, "[待补充]"))
        add_border(ws[f'A{row}'])
        add_border(ws[f'B{row}'])
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    row += 1
    
    # 二、金融机构
    style_header(ws[f'A{row}'], "二、金融机构信息")
    ws.merge_cells(f'A{row}:E{row}')
    ws.row_dimensions[row].height = 25
    row += 1
    
    for field, key in [("金融机构名称", "financial_institution"), 
                       ("金融机构ID", "financial_institution_id")]:
        ws[f'A{row}'] = field
        ws[f'B{row}'] = info.get(key, "[待补充]")
        style_cell(ws[f'A{row}'], field, bold=True, fill_color=colors["light_gray"])
        style_cell(ws[f'B{row}'], info.get(key, "[待补充]"))
        add_border(ws[f'A{row}'])
        add_border(ws[f'B{row}'])
        ws.merge_cells(f'B{row}:E{row}')
        row += 1
    
    row += 1
    
    # 三、账户余额
    style_header(ws[f'A{row}'], "三、账户余额/价值")
    ws.merge_cells(f'A{row}:E{row}')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # 表头
    for col, h in enumerate(["资产类别", "币种", "期末余额", "备注"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style_subheader(cell, h)
        add_border(cell)
    
    # 资产数据
    assets = [
        ("股票和股票期权", "HKD", info.get("stock_value", "[待补充]"), ""),
        ("基金", "HKD", info.get("fund_value", "[待补充]"), ""),
        ("现金", "HKD", info.get("cash_value", "[待补充]"), ""),
        ("合计", "HKD", info.get("total_value", "[待补充]"), "资产净值"),
    ]
    
    for asset, currency, value, note in assets:
        row += 1
        ws[f'A{row}'] = asset
        ws[f'B{row}'] = currency
        ws[f'C{row}'] = value
        ws[f'D{row}'] = note
        
        if "合计" in asset:
            for col in range(1, 5):
                style_cell(ws.cell(row=row, column=col), ws.cell(row=row, column=col).value, bold=True, fill_color=colors["highlight"])
                add_border(ws.cell(row=row, column=col))
        else:
            for col in range(1, 5):
                style_cell(ws.cell(row=row, column=col), ws.cell(row=row, column=col).value)
                add_border(ws.cell(row=row, column=col))
    
    row += 2
    
    # 四、交易记录
    style_header(ws[f'A{row}'], "四、交易记录凭证")
    ws.merge_cells(f'A{row}:E{row}')
    ws.row_dimensions[row].height = 25
    row += 1
    
    # 表头
    for col, h in enumerate(["日期", "交易类型", "产品名称", "数量", "金额"], 1):
        cell = ws.cell(row=row, column=col, value=h)
        style_subheader(cell, h)
        add_border(cell)
    
    # 交易记录
    for tx in info.get("transactions", []):
        row += 1
        ws[f'A{row}'] = tx.get("date", "")
        ws[f'B{row}'] = tx.get("type", "")
        ws[f'C{row}'] = tx.get("name", "")
        ws[f'D{row}'] = tx.get("quantity", "")
        ws[f'E{row}'] = tx.get("amount", "")
        for col in range(1, 6):
            style_cell(ws.cell(row=row, column=col), ws.cell(row=row, column=col).value)
            add_border(ws.cell(row=row, column=col))
    
    # 列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    wb.save(output_path)
    return output_path

# ============== 主函数 ==============
def main():
    if len(sys.argv) < 2:
        print("""
CRS 申报表生成器 v2.0
用法: python crs_generator.py <pdf文件> [输出文件]
示例: python crs_generator.py client_statement.pdf CRS_report.xlsx
        """)
        # 生成模板
        generate_excel(DEFAULT_INFO, "CRS_申报表_模板.xlsx")
        print("已生成模板: CRS_申报表_模板.xlsx")
        return
    
    pdf_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else "CRS_申报表.xlsx"
    
    print(f"📄 处理文件: {pdf_path}")
    
    # 解析PDF
    info = parse_pdf(pdf_path)
    if info:
        print("✅ PDF解析完成")
        print(f"   提取到 {len(info.get('transactions', []))} 条交易记录")
    else:
        print("⚠️ PDF解析失败，使用模板")
        info = DEFAULT_INFO
    
    # 生成Excel
    output = generate_excel(info, output_path)
    print(f"✅ CRS申报表已生成: {output}")

if __name__ == "__main__":
    main()