import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
import argparse
import sys

def export_to_excel(input_csv, output_xlsx, sheet_name):
    print(f"📊 Exporting {input_csv} to Excel: {output_xlsx}...")
    try:
        df = pd.read_csv(input_csv)
    except Exception as e:
        print(f"❌ Error reading input CSV: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    print("📝 Writing data...")
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Basic formatting
    print("🎨 Applying basic formatting...")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    print(f"💾 Saving workbook to {output_xlsx}...")
    wb.save(output_xlsx)
    print("✅ Export complete.")

def main():
    parser = argparse.ArgumentParser(description="Export CSV data to formatted Excel.")
    parser.add_argument("input", help="Input CSV file")
    parser.add_argument("output", help="Output Excel file (.xlsx)")
    parser.add_argument("--sheet", default="Data", help="Sheet name")
    
    args = parser.parse_args()
    export_to_excel(args.input, args.output, args.sheet)

if __name__ == "__main__":
    main()
