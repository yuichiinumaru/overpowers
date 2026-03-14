#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
财务报表解析脚本
功能：解析 Excel 格式的财务报表（XLS/XLSX），提取关键财务指标
"""

import argparse
import json
import sys
from typing import Dict, Any, List
import openpyxl
from datetime import datetime


def parse_financial_statement(file_path: str) -> Dict[str, Any]:
    """
    解析财务报表文件
    
    参数:
        file_path: 财务报表文件路径（XLS/XLSX）
    
    返回:
        包含财务数据的字典
    """
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        result = {
            'file_name': file_path.split('/')[-1],
            'parse_time': datetime.now().isoformat(),
            'sheets': {},
            'summary': {}
        }
        
        # 解析每个工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            
            # 读取所有数据
            for row in sheet.iter_rows(values_only=True):
                # 过滤掉完全空行
                if any(cell is not None for cell in row):
                    sheet_data.append([str(cell) if cell is not None else '' for cell in row])
            
            result['sheets'][sheet_name] = sheet_data
        
        # 提取关键财务指标（智能辅助）
        result['summary'] = extract_key_metrics(result)
        
        return result
        
    except Exception as e:
        return {
            'error': f'解析失败: {str(e)}',
            'file_path': file_path
        }


def extract_key_metrics(data: Dict[str, Any]) -> Dict[str, Any]:
    """
    从解析的数据中提取关键财务指标
    
    参数:
        data: 解析后的完整数据
    
    返回:
        关键指标字典
    """
    metrics = {
        'revenue': None,
        'gross_profit': None,
        'net_profit': None,
        'total_assets': None,
        'total_liabilities': None,
        'cash_flow': None,
        'detected_tables': []
    }
    
    # 遍历所有工作表，尝试识别关键指标
    for sheet_name, sheet_data in data['sheets'].items():
        for row_idx, row in enumerate(sheet_data):
            row_text = ' '.join(row).lower()
            
            # 识别收入相关
            if any(keyword in row_text for keyword in ['营业收入', '总收入', '收入', 'revenue']):
                if len(row) > 1 and row[1]:
                    try:
                        metrics['revenue'] = float(row[1])
                    except (ValueError, IndexError):
                        pass
            
            # 识别利润相关
            if any(keyword in row_text for keyword in ['净利润', '净利润', 'net profit']):
                if len(row) > 1 and row[1]:
                    try:
                        metrics['net_profit'] = float(row[1])
                    except (ValueError, IndexError):
                        pass
            
            # 识别总资产
            if any(keyword in row_text for keyword in ['资产总计', '总资产', 'total assets']):
                if len(row) > 1 and row[1]:
                    try:
                        metrics['total_assets'] = float(row[1])
                    except (ValueError, IndexError):
                        pass
            
            # 识别总负债
            if any(keyword in row_text for keyword in ['负债合计', '总负债', 'total liabilities']):
                if len(row) > 1 and row[1]:
                    try:
                        metrics['total_liabilities'] = float(row[1])
                    except (ValueError, IndexError):
                        pass
    
    # 检测到的表格
    metrics['detected_tables'] = list(data['sheets'].keys())
    
    return metrics


def main():
    parser = argparse.ArgumentParser(description='解析财务报表文件')
    parser.add_argument('--file', required=True, help='财务报表文件路径（XLS/XLSX）')
    parser.add_argument('--output', help='输出JSON文件路径（可选，默认打印到标准输出）')
    
    args = parser.parse_args()
    
    # 解析文件
    result = parse_financial_statement(args.file)
    
    # 输出结果
    if 'error' in result:
        print(json.dumps({'error': result['error']}, ensure_ascii=False, indent=2), file=sys.stderr)
        sys.exit(1)
    
    output_json = json.dumps(result, ensure_ascii=False, indent=2)
    
    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(output_json)
        print(f'结果已保存到: {args.output}')
    else:
        print(output_json)


if __name__ == '__main__':
    main()
